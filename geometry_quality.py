# Rotate and normalize your generated mesh to align with gt mesh (*_align.obj)
# Copy them to generated_* dir for comparison
# Calculate geometry quality (Chamfer distance and F-score)
import os
import xlwt
from metrics import eval_pointcloud
import torch
import trimesh
#from rotate_glb import uniform_scale_meshes, normalize_mesh, rotate_mesh_y
import shutil
import numpy as np
from openpyxl import Workbook
import argparse


def fix_vert_color_glb(mesh_path):
    from pygltflib import GLTF2, Material, PbrMetallicRoughness
    obj1 = GLTF2().load(mesh_path)
    obj1.meshes[0].primitives[0].material = 0
    obj1.materials.append(Material(
        pbrMetallicRoughness = PbrMetallicRoughness(
            baseColorFactor = [1.0, 1.0, 1.0, 1.0],
            metallicFactor = 0.,
            roughnessFactor = 1.0,
        ),
        emissiveFactor = [0.0, 0.0, 0.0],
        doubleSided = True,
    ))
    obj1.save(mesh_path)
    
# def copy_glb_files(original_dir, target_dir):
#     os.makedirs(target_dir, exist_ok=True)
#     for folder in os.listdir(original_dir):
#         folder_path = os.path.join(original_dir, folder)
#         if os.path.isdir(folder_path):
#             prefix = "_".join(folder.split('_')[:-3])
#             azim = folder.split('_')[-1].replace("azim", "")
#             for file in os.listdir(folder_path):
#                 # Choose your generated format: .glb, .obj, .etc
#                 if  file.endswith(".glb"): 
#                     source_file = os.path.join(folder_path, file)
#                     target_folder = os.path.join(target_dir, prefix)
#                     os.makedirs(target_folder, exist_ok=True)
#                     # Choose your generated format: .glb, .obj, .etc
#                     target_file = os.path.join(target_folder, f"{prefix}.glb")
#                     scene = trimesh.load(source_file, force='scene')  
#                     if isinstance(scene, trimesh.Scene):
#                         mesh = next(iter(scene.geometry.values()))
#                     else:
#                         mesh = scene
#                     mesh = rotate_mesh_y(mesh, int(azim))
#                     mesh.export(target_file)
#                     fix_vert_color_glb(target_file)


def find_glb_files(rotated_dir):
    glb_files = []
    for root, dirs, files in os.walk(rotated_dir):
        for file in files:
            if file.lower().endswith('.glb') and 'icp' in file.lower():
                subdir = os.path.basename(root)
                glb_path = os.path.join(root, file)
                # 保存子目录名和文件名，以便后续匹配和结果记录
                glb_files.append((subdir, file, glb_path))
    return glb_files

def find_gt_obj(original_dir, subdir):
    # 更新为新的文件路径结构
    obj_filename = f"{subdir}_align.obj"
    obj_path = os.path.join(original_dir, subdir, obj_filename)
    if os.path.isfile(obj_path):
        return obj_path
    else:
        return None

def write_to_excel(results, output_excel):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Evaluation Results'

    headers = [
        'Subdir',
        'Prediction File',
        'Chamfer Distance',
        'F-Score (0.005)',
        'F-Score (0.03)',
        'F-Score (0.1)',
        'F-Score (0.2)',
        'F-Score (0.5)'
    ]

    # 写入表头
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    # 按子目录组织结果
    grouped_results = {}
    for result in results:
        subdir = result['Subdir']
        if subdir not in grouped_results:
            grouped_results[subdir] = []
        grouped_results[subdir].append(result)

    row = 2  # 从第2行开始写入数据
    for subdir, subdir_results in grouped_results.items():
        # 写入每个子目录的所有预测结果
        for result in subdir_results:
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=row, column=col, value=result.get(header, ''))
            row += 1
        
        # 计算并写入该子目录的平均值
        avg_values = {header: 0.0 for header in headers[2:]}  # 跳过Subdir和Prediction File
        for result in subdir_results:
            for header in headers[2:]:
                avg_values[header] += result.get(header, 0)
        
        for header in avg_values:
            avg_values[header] /= len(subdir_results)
        
        sheet.cell(row=row, column=1, value=f"{subdir} Average")
        for col, header in enumerate(headers[2:], start=3):
            sheet.cell(row=row, column=col, value=avg_values[header])
        row += 1
        row += 1  # 添加空行分隔不同子目录的结果

    # 计算所有结果的总平均值
    total_results = len(results)
    total_avg = {header: 0.0 for header in headers[2:]}
    for result in results:
        for header in headers[2:]:
            total_avg[header] += result.get(header, 0)
    
    for header in total_avg:
        total_avg[header] /= total_results

    sheet.cell(row=row, column=1, value="Total Average")
    for col, header in enumerate(headers[2:], start=3):
        sheet.cell(row=row, column=col, value=total_avg[header])

    # 直接使用原始文件名保存
    workbook.save(output_excel)
    print(f"Results successfully saved to {output_excel}")

def analyze_best_results(excel_file):
    from openpyxl import load_workbook
    
    print("\nAnalyzing best F-Score (0.03) results for each subdirectory...")
    
    # 加载Excel文件
    wb = load_workbook(excel_file)
    sheet = wb.active
    
    # 获取表头
    headers = [cell.value for cell in sheet[1]]
    
    # 创建数据结构来存储每个子目录的结果
    subdir_results = {}
    
    # 读取数据
    current_row = 2
    while current_row < sheet.max_row:
        cell_value = sheet.cell(row=current_row, column=1).value
        if cell_value and not cell_value.endswith('Average') and cell_value != 'Total Average':
            subdir = sheet.cell(row=current_row, column=1).value
            pred_file = sheet.cell(row=current_row, column=2).value
            
            # 获取所有指标
            metrics = {
                'Prediction File': pred_file,
                'Chamfer Distance': sheet.cell(row=current_row, column=3).value,
                'F-Score (0.005)': sheet.cell(row=current_row, column=4).value,
                'F-Score (0.03)': sheet.cell(row=current_row, column=5).value,
                'F-Score (0.1)': sheet.cell(row=current_row, column=6).value,
                'F-Score (0.2)': sheet.cell(row=current_row, column=7).value,
                'F-Score (0.5)': sheet.cell(row=current_row, column=8).value
            }
            
            if subdir not in subdir_results:
                subdir_results[subdir] = {
                    'best_fscore_03': float('-inf'),
                    'best_metrics': None
                }
            
            # 更新最佳F-Score (0.03)及其对应的所有指标
            if metrics['F-Score (0.03)'] > subdir_results[subdir]['best_fscore_03']:
                subdir_results[subdir]['best_fscore_03'] = metrics['F-Score (0.03)']
                subdir_results[subdir]['best_metrics'] = metrics
        
        current_row += 1
    
    # 创建新的Excel文件保存最佳结果
    best_results_wb = Workbook()
    best_sheet = best_results_wb.active
    best_sheet.title = 'Best F-Score Results'
    
    # 写入表头
    headers = ['Subdir', 'Prediction File', 'Chamfer Distance', 
              'F-Score (0.005)', 'F-Score (0.03)', 'F-Score (0.1)',
              'F-Score (0.2)', 'F-Score (0.5)']
    for col, header in enumerate(headers, start=1):
        best_sheet.cell(row=1, column=col, value=header)
    
    # 计算平均值
    avg_metrics = {
        'Chamfer Distance': 0.0,
        'F-Score (0.005)': 0.0,
        'F-Score (0.03)': 0.0,
        'F-Score (0.1)': 0.0,
        'F-Score (0.2)': 0.0,
        'F-Score (0.5)': 0.0
    }
    
    # 写入每个子目录的最佳结果
    row = 2
    for subdir, results in subdir_results.items():
        metrics = results['best_metrics']
        best_sheet.cell(row=row, column=1, value=subdir)
        best_sheet.cell(row=row, column=2, value=metrics['Prediction File'])
        best_sheet.cell(row=row, column=3, value=metrics['Chamfer Distance'])
        best_sheet.cell(row=row, column=4, value=metrics['F-Score (0.005)'])
        best_sheet.cell(row=row, column=5, value=metrics['F-Score (0.03)'])
        best_sheet.cell(row=row, column=6, value=metrics['F-Score (0.1)'])
        best_sheet.cell(row=row, column=7, value=metrics['F-Score (0.2)'])
        best_sheet.cell(row=row, column=8, value=metrics['F-Score (0.5)'])
        
        # 累加各指标值用于计算平均值
        for metric in avg_metrics:
            avg_metrics[metric] += metrics[metric]
        
        row += 1
    
    # 计算并写入平均值
    num_subdirs = len(subdir_results)
    for metric in avg_metrics:
        avg_metrics[metric] /= num_subdirs
    
    best_sheet.cell(row=row, column=1, value="Average")
    best_sheet.cell(row=row, column=3, value=avg_metrics['Chamfer Distance'])
    best_sheet.cell(row=row, column=4, value=avg_metrics['F-Score (0.005)'])
    best_sheet.cell(row=row, column=5, value=avg_metrics['F-Score (0.03)'])
    best_sheet.cell(row=row, column=6, value=avg_metrics['F-Score (0.1)'])
    best_sheet.cell(row=row, column=7, value=avg_metrics['F-Score (0.2)'])
    best_sheet.cell(row=row, column=8, value=avg_metrics['F-Score (0.5)'])
    
    # 保存结果
    output_file = excel_file.replace('.xlsx', '_best_fscore.xlsx')
    best_results_wb.save(output_file)
    print(f"\nBest F-Score (0.03) results saved to: {output_file}")
    
    # 打印结果
    print("\nBest F-Score (0.03) results for each subdirectory:")
    print("-" * 80)
    for subdir, results in subdir_results.items():
        metrics = results['best_metrics']
        print(f"\nSubdirectory: {subdir}")
        print(f"Best F-Score (0.03): {metrics['F-Score (0.03)']:.6f}")
        print(f"File: {metrics['Prediction File']}")
        print(f"Chamfer Distance: {metrics['Chamfer Distance']:.6f}")
        print(f"Other F-Scores: 0.005={metrics['F-Score (0.005)']:.6f}, "
              f"0.1={metrics['F-Score (0.1)']:.6f}, "
              f"0.2={metrics['F-Score (0.2)']:.6f}, "
              f"0.5={metrics['F-Score (0.5)']:.6f}")
    
    print("\nAverage across all subdirectories:")
    print(f"Chamfer Distance: {avg_metrics['Chamfer Distance']:.6f}")
    print(f"F-Score (0.005): {avg_metrics['F-Score (0.005)']:.6f}")
    print(f"F-Score (0.03): {avg_metrics['F-Score (0.03)']:.6f}")
    print(f"F-Score (0.1): {avg_metrics['F-Score (0.1)']:.6f}")
    print(f"F-Score (0.2): {avg_metrics['F-Score (0.2)']:.6f}")
    print(f"F-Score (0.5): {avg_metrics['F-Score (0.5)']:.6f}")
    print("-" * 80)

def main():
    parser = argparse.ArgumentParser(description='评估3D网格模型的几何质量')
    parser.add_argument('--rotated-dir', default='trellis_rotated', 
                      help='预测的glb文件目录 (默认: trellis_rotated)')
    parser.add_argument('--original-dir', default='aligned_gt',
                      help='GT mesh目录 (默认: aligned_gt)')
    parser.add_argument('--output-excel', default=None,
                      help='输出Excel文件路径 (默认: 根据rotated-dir自动生成)')
    parser.add_argument('--sample-points', type=int, default=200000,
                      help='采样点数量 (默认: 200000)')
    parser.add_argument('--eval-type', default='real_obj',
                      help='评估类型 (默认: real_obj)')
    parser.add_argument('--analyze-only', action='store_true',
                      help='仅分析已有的Excel文件，不进行新的评估')
    
    args = parser.parse_args()
    
    # 如果未指定输出Excel文件，则根据rotated_dir自动生成
    if args.output_excel is None:
        # 提取rotated_dir的前缀作为输出文件名前缀
        prefix = args.rotated_dir.split('_')[0] if '_' in args.rotated_dir else args.rotated_dir
        args.output_excel = f"{prefix}_geometry_quality.xlsx"  # 直接使用 .xlsx 扩展名
    
    print(f"\n几何质量评估配置:")
    print(f"预测模型目录: {args.rotated_dir}")
    print(f"参考模型目录: {args.original_dir}")
    print(f"输出Excel文件: {args.output_excel}")
    print(f"采样点数量: {args.sample_points}")
    print(f"评估类型: {args.eval_type}")
    
    # 如果只分析已有Excel文件
    if args.analyze_only:
        if os.path.exists(args.output_excel):
            print(f"\n仅分析已有Excel文件: {args.output_excel}")
            analyze_best_results(args.output_excel)
        else:
            print(f"错误: 指定的Excel文件不存在: {args.output_excel}")
        return
    
    # 查找所有glb文件
    glb_files = find_glb_files(args.rotated_dir)
    if not glb_files:
        print(f"错误: 在 {args.rotated_dir} 中未找到.glb文件")
        return
    
    print(f"\n找到 {len(glb_files)} 个预测模型文件")
    
    results = []
    #processed_gt_paths = set()  # 用于记录已处理过的GT mesh路径
    
    for subdir, pred_file, glb_path in glb_files:
        gt_obj_path = find_gt_obj(args.original_dir, subdir)
        if not gt_obj_path:
            print(f"警告: 未找到对应的参考模型: {subdir}. 跳过.")
            continue
        
        print(f"\n处理: {subdir}/{pred_file}")
        print(f'参考模型路径: {gt_obj_path}')
        print(f'预测模型路径: {glb_path}')
        
        # 预处理预测的mesh
        try:
            fix_vert_color_glb(glb_path)
            print(f"成功处理预测模型: {glb_path}")
        except Exception as e:
            print(f"警告: 处理预测模型失败 {glb_path}: {str(e)}")
        
        # 只处理还未处理过的GT mesh
        # if gt_obj_path not in processed_gt_paths:
        #     try:
        #         fix_vert_color_glb(gt_obj_path)
        #         processed_gt_paths.add(gt_obj_path)
        #         print(f"成功处理参考模型: {gt_obj_path}")
        #     except Exception as e:
        #         print(f"参考模型已处理 {gt_obj_path}: {str(e)}")
        
        # 计算指标
        print(f"计算几何质量指标...")
        metrics = eval_pointcloud(
            pre_mesh_file=glb_path,
            gt_mesh_file=gt_obj_path,
            samplepoint=args.sample_points,
            eval_type=args.eval_type
        )
        
        # 记录结果
        results.append({
            'Subdir': subdir,
            'Prediction File': pred_file,
            'Chamfer Distance': metrics.get('chamfer-L2', 0),
            'F-Score (0.005)': metrics.get('f-score-005', 0),
            'F-Score (0.03)': metrics.get('f-score-03', 0),
            'F-Score (0.1)': metrics.get('f-score-1', 0),
            'F-Score (0.2)': metrics.get('f-score-2', 0),
            'F-Score (0.5)': metrics.get('f-score-5', 0)
        })
        
        # 打印当前结果
        print(f"Chamfer Distance: {metrics.get('chamfer-L2', 0):.6f}")
        print(f"F-Score (0.005): {metrics.get('f-score-005', 0):.6f}")
        print(f"F-Score (0.03): {metrics.get('f-score-03', 0):.6f}")
        print(f"F-Score (0.1): {metrics.get('f-score-1', 0):.6f}")
        print(f"F-Score (0.2): {metrics.get('f-score-2', 0):.6f}")
        print(f"F-Score (0.5): {metrics.get('f-score-5', 0):.6f}")
    
    # 写入Excel
    write_to_excel(results, args.output_excel)
    
    # 分析最佳结果
    analyze_best_results(args.output_excel)

if __name__ == "__main__":
    main()
